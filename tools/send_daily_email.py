"""
Send a daily digest email with applications.xlsx attached.

The email includes:
- Upcoming interviews / assessments from Google Calendar (next 14 days)
- A Claude-generated priority summary
- The full applications.xlsx as an attachment

Run directly:  python tools/send_daily_email.py
"""

import base64
import os
import sys
from datetime import datetime, timedelta, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from dotenv import load_dotenv

load_dotenv()

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from anthropic import Anthropic
from googleapiclient.errors import HttpError
from tools.google_auth import get_gmail_send_service, get_calendar_service
from config import CALENDAR_ID, TIMEZONE, OUTPUT_XLSX
import openpyxl

RECIPIENT = "ryanliu61799@gmail.com"
LOOKAHEAD_DAYS = 14

client = Anthropic()


def load_applications() -> list[dict]:
    """Load all records from applications.xlsx."""
    path = os.path.join(PROJECT_ROOT, OUTPUT_XLSX)
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path)
    if "Applications" not in wb.sheetnames:
        return []

    ws = wb["Applications"]
    headers = [cell.value for cell in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        records.append(dict(zip(headers, row)))
    return records


def get_upcoming_events() -> list[dict]:
    """Fetch calendar events for the next LOOKAHEAD_DAYS days."""
    try:
        service = get_calendar_service()
        now = datetime.now(timezone.utc)
        time_min = now.isoformat()
        time_max = (now + timedelta(days=LOOKAHEAD_DAYS)).isoformat()

        result = service.events().list(
            calendarId=CALENDAR_ID,
            timeMin=time_min,
            timeMax=time_max,
            singleEvents=True,
            orderBy="startTime",
        ).execute()

        events = []
        for e in result.get("items", []):
            start = e.get("start", {})
            dt = start.get("dateTime") or start.get("date", "")
            events.append({
                "summary": e.get("summary", ""),
                "start": dt,
                "description": e.get("description", ""),
            })
        return events
    except HttpError as e:
        print(f"⚠️  Calendar fetch failed: {e}")
        return []


def build_summary(records: list[dict], events: list[dict]) -> str:
    """Use Claude to generate a prioritised briefing."""
    today = datetime.now().strftime("%Y-%m-%d")

    # Condense records for the prompt — only active ones
    active_statuses = {
        "Online Assessment", "First Round Interview",
        "Second Round Interview", "Assessment Centre", "Offer", "Applied", "Pending"
    }
    active = [r for r in records if r.get("Status") in active_statuses]

    records_text = "\n".join(
        f"- {r.get('Company')} | {r.get('Role')} | {r.get('Status')} | "
        f"Next Step: {r.get('Next Step') or 'N/A'} | Next Deadline: {r.get('Next Deadline') or 'N/A'}"
        for r in active
    ) or "No active applications."

    events_text = "\n".join(
        f"- {e['start']}: {e['summary']}"
        for e in events
    ) or "No upcoming calendar events."

    prompt = f"""Today is {today}. You are helping a finance student track their internship applications.

ACTIVE APPLICATIONS:
{records_text}

UPCOMING CALENDAR EVENTS (next {LOOKAHEAD_DAYS} days):
{events_text}

Write a concise daily briefing (plain text, no markdown) covering:
1. Any imminent deadlines or events (in the next 7 days) — include exact dates and times if known
2. What the student should focus on TODAY
3. A short status summary (e.g. X active, Y upcoming interviews)

Be direct and actionable. Max 200 words."""

    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text.strip()


def build_email(summary: str, xlsx_path: str) -> MIMEMultipart:
    """Construct the MIME email with summary body and xlsx attachment."""
    today_str = datetime.now().strftime("%A, %d %B %Y")

    msg = MIMEMultipart()
    msg["To"] = RECIPIENT
    msg["From"] = RECIPIENT
    msg["Subject"] = f"Internship Tracker — {today_str}"

    body = f"Daily Internship Briefing\n{'='*40}\n\n{summary}\n\n---\nFull tracker attached."
    msg.attach(MIMEText(body, "plain"))

    if os.path.exists(xlsx_path):
        with open(xlsx_path, "rb") as f:
            attachment = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            attachment.add_header("Content-Disposition", "attachment", filename="applications.xlsx")
            msg.attach(attachment)

    return msg


def send_email(msg: MIMEMultipart):
    """Send the email via Gmail API."""
    service = get_gmail_send_service()
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    print(f"✅ Email sent to {RECIPIENT}")


def main():
    print("📊 Loading applications...")
    records = load_applications()
    print(f"   {len(records)} records loaded")

    print("📅 Fetching upcoming calendar events...")
    events = get_upcoming_events()
    print(f"   {len(events)} events in next {LOOKAHEAD_DAYS} days")

    print("🤖 Generating summary...")
    summary = build_summary(records, events)
    print(f"\n--- SUMMARY PREVIEW ---\n{summary}\n-----------------------\n")

    xlsx_path = os.path.join(PROJECT_ROOT, OUTPUT_XLSX)
    msg = build_email(summary, xlsx_path)

    print("📨 Sending email...")
    send_email(msg)


if __name__ == "__main__":
    main()
