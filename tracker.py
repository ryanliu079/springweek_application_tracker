# tracker.py
import argparse
import json
import os
import re
from datetime import datetime

import base64

from dotenv import load_dotenv

load_dotenv()

from anthropic import Anthropic
from googleapiclient.errors import HttpError
from tools.google_auth import get_gmail_service, get_calendar_service
from config import (
    CALENDAR_ID,
    DATE_CUTOFF,
    MAX_THREADS,
    OUTPUT_XLSX,
    ROLE_TYPES,
    SEARCH_QUERY,
    SHEET_NAME,
    STATUSES,
    TIMEZONE,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

client = Anthropic()

# ─────────────────────────────────────────
# STEP 1: FETCH EMAILS VIA GMAIL MCP
# ─────────────────────────────────────────

def _get_header(headers: list, name: str) -> str:
    """Extract a header value by name from a Gmail message headers list."""
    for h in headers:
        if h["name"].lower() == name.lower():
            return h["value"]
    return ""

def _decode_body(payload: dict) -> str:
    """Recursively extract plain-text body from a Gmail message payload."""
    mime = payload.get("mimeType", "")
    if mime == "text/plain":
        data = payload.get("body", {}).get("data", "")
        return base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace") if data else ""
    if mime.startswith("multipart/"):
        for part in payload.get("parts", []):
            text = _decode_body(part)
            if text:
                return text
    return ""

def fetch_application_emails(date_cutoff: str = DATE_CUTOFF) -> list[dict]:
    """Fetch relevant email threads from ryanliu61799@gmail.com via Gmail API."""
    print("📥 Fetching application emails from Gmail...")
    service = get_gmail_service()

    query = f"({SEARCH_QUERY.strip()}) after:{date_cutoff.replace('/', '-')}"
    thread_ids = []
    page_token = None

    while len(thread_ids) < MAX_THREADS:
        kwargs = {"userId": "me", "q": query, "maxResults": min(500, MAX_THREADS - len(thread_ids))}
        if page_token:
            kwargs["pageToken"] = page_token
        try:
            result = service.users().threads().list(**kwargs).execute()
        except HttpError as e:
            print(f"   ⚠️  Gmail API error: {e}")
            break

        thread_ids.extend(t["id"] for t in result.get("threads", []))
        page_token = result.get("nextPageToken")
        if not page_token:
            break

    print(f"   Found {len(thread_ids)} matching threads — fetching details...")
    threads = []
    for tid in thread_ids[:MAX_THREADS]:
        try:
            t = service.users().threads().get(userId="me", id=tid, format="full").execute()
            msgs = t.get("messages", [])
            if not msgs:
                continue

            all_subjects = []
            latest_msg = None
            latest_date = ""
            for msg in msgs:
                hdrs = msg.get("payload", {}).get("headers", [])
                subj = _get_header(hdrs, "Subject")
                if subj:
                    all_subjects.append(subj)
                date_str = _get_header(hdrs, "Date")
                if date_str > latest_date:
                    latest_date = date_str
                    latest_msg = msg

            if not latest_msg:
                latest_msg = msgs[-1]

            payload = latest_msg.get("payload", {})
            hdrs = payload.get("headers", [])
            body = _decode_body(payload)[:500]

            threads.append({
                "thread_id": tid,
                "subject": _get_header(hdrs, "Subject"),
                "from": _get_header(hdrs, "From"),
                "date": _get_header(hdrs, "Date"),
                "body_snippet": body or latest_msg.get("snippet", "")[:500],
                "all_subjects": all_subjects,
            })
        except HttpError as e:
            print(f"   ⚠️  Skipped thread {tid}: {e}")

    return threads


# ─────────────────────────────────────────
# STEP 2: PARSE EMAILS INTO STRUCTURED DATA
# ─────────────────────────────────────────

PARSE_SYSTEM = """Parse internship application email threads into structured records.
Return ONLY a JSON array (one object per thread) with keys:
  company, role, role_type, division, location,
  date_applied (ISO date), deadline (ISO date or null),
  status (Applied|Online Assessment|First Round Interview|Second Round Interview|Assessment Centre|Offer|Rejected|Withdrawn|Pending),
  last_updated (ISO date), next_step (string or null),
  next_deadline (ISO date or null), notes, thread_id
If unsure, use "Pending" for status and null for unknowns.
role_type: Spring Week|Summer Internship|Off-Cycle Internship|Graduate Scheme|Part-Time / Other

SKIP these — return null for threads that are clearly NOT real job applications:
- LinkedIn notifications (job alerts, profile views, skill endorsements)
- Marketing/promotional emails (e.g. AmplifyME, Forage course ads)
- Generic job board listings with no confirmation the user actually applied
- Account registration emails with no specific role (unless part of a known application flow)
Return null in the array slot for skipped threads (they will be filtered out).

Rejection inference: Today's date is {today}. If the most recent email in a thread is an application confirmation (no follow-up from the company) and the email is more than 6 weeks old, infer status "Rejected" — finance firms do not leave candidates waiting that long without a response. Add a note: "Inferred rejection — no response after 6+ weeks." """

BATCH_SIZE = 10

def _format_thread(t: dict) -> str:
    return (f"thread_id:{t.get('thread_id')} | from:{t.get('from')} | date:{t.get('date')}\n"
            f"subjects:{t.get('all_subjects', [])} | snippet:{t.get('body_snippet', '')[:300]}")

def parse_threads_batch(threads: list[dict]) -> list[dict]:
    """Parse a batch of threads in one API call. Falls back to Sonnet on JSON error."""
    system = PARSE_SYSTEM.format(today=datetime.now().strftime("%Y-%m-%d"))
    prompt = "Parse each thread below:\n\n" + "\n---\n".join(
        f"[{i+1}] {_format_thread(t)}" for i, t in enumerate(threads)
    )

    for model in ("claude-haiku-4-5-20251001", "claude-sonnet-4-6"):
        response = client.messages.create(
            model=model,
            max_tokens=600 * len(threads),
            system=system,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response.content[0].text
        try:
            match = re.search(r'\[.*\]', raw, re.DOTALL)
            records = json.loads(match.group() if match else raw)
            if isinstance(records, list):
                break
        except (json.JSONDecodeError, AttributeError):
            if model == "claude-sonnet-4-6":
                records = []

    # Merge thread_ids from source since model may omit them; filter nulls (skipped threads)
    result = []
    for i, rec in enumerate(records):
        if not rec:
            continue
        if i < len(threads):
            rec.setdefault("thread_id", threads[i].get("thread_id", ""))
        rec["email_thread_id"] = rec.pop("thread_id", "")
        rec["calendar_event_id"] = ""
        result.append(rec)
    return result

def parse_all_threads(threads: list[dict]) -> list[dict]:
    """Batch-parse all threads, BATCH_SIZE at a time."""
    records = []
    for i in range(0, len(threads), BATCH_SIZE):
        batch = threads[i:i + BATCH_SIZE]
        print(f"   Parsing threads {i+1}–{min(i+BATCH_SIZE, len(threads))} of {len(threads)}...")
        try:
            records.extend(parse_threads_batch(batch))
        except Exception as e:
            print(f"   ⚠️  Batch {i//BATCH_SIZE + 1} failed: {e}")
    return records


# ─────────────────────────────────────────
# STEP 3: DEDUPLICATE RECORDS
# ─────────────────────────────────────────

def _normalize_role(role: str) -> str:
    """Strip year suffixes, job codes, and filler words for fuzzy matching."""
    role = (role or "").lower().strip()
    role = re.sub(r'\b20\d{2}\b', '', role)          # strip years like 2025, 2026
    role = re.sub(r'\(jr-[\w-]+\)', '', role)         # strip job codes like (JR-0000082018)
    role = re.sub(r'\b(programme|program|the|and|&)\b', '', role)
    role = re.sub(r'\s+', ' ', role).strip()
    return role

def deduplicate(records: list[dict]) -> list[dict]:
    """Merge multiple records for the same company+role, keeping latest status.
    Uses company name alone as secondary key to catch model role-name variance."""
    # First pass: exact company+normalised-role key
    seen = {}
    for r in records:
        key = ((r.get("company") or "").lower().strip(), _normalize_role(r.get("role")))
        if key not in seen:
            seen[key] = r
        else:
            existing = seen[key]
            try:
                r_idx = STATUSES.index(r.get("status", "Pending"))
                e_idx = STATUSES.index(existing.get("status", "Pending"))
            except ValueError:
                r_idx = e_idx = 0
            if r_idx > e_idx:
                r["date_applied"] = existing.get("date_applied")
                seen[key] = r

    # Second pass: collapse remaining duplicates by company name alone
    by_company = {}
    for r in seen.values():
        co = (r.get("company") or "").lower().strip()
        if co not in by_company:
            by_company[co] = r
        else:
            existing = by_company[co]
            try:
                r_idx = STATUSES.index(r.get("status", "Pending"))
                e_idx = STATUSES.index(existing.get("status", "Pending"))
            except ValueError:
                r_idx = e_idx = 0
            if r_idx > e_idx:
                r["date_applied"] = existing.get("date_applied")
                by_company[co] = r
    return list(by_company.values())


# ─────────────────────────────────────────
# STEP 4: WRITE XLSX
# ─────────────────────────────────────────

STATUS_COLOURS = {
    "Applied":                  "D6E4F0",
    "Online Assessment":        "FFF3CD",
    "First Round Interview":    "D4EDDA",
    "Second Round Interview":   "A8D5BA",
    "Assessment Centre":        "7EC8A4",
    "Offer":                    "28A745",
    "Rejected":                 "F8D7DA",
    "Withdrawn":                "E2E3E5",
    "Pending":                  "FFFFFF",
}

HEADERS = [
    "Company", "Role", "Role Type", "Division", "Location",
    "Date Applied", "Deadline", "Status", "Last Updated",
    "Next Step", "Next Deadline", "Notes", "Thread ID", "Calendar Event ID"
]

RECORD_KEYS = [
    "company", "role", "role_type", "division", "location",
    "date_applied", "deadline", "status", "last_updated",
    "next_step", "next_deadline", "notes", "email_thread_id", "calendar_event_id"
]

def write_xlsx(records: list[dict], path: str = OUTPUT_XLSX):
    # Load existing workbook if updating, otherwise create fresh
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME, 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # Header row
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for row, rec in enumerate(records, 2):
        status = rec.get("status", "Pending")
        row_colour = STATUS_COLOURS.get(status, "FFFFFF")
        fill = PatternFill("solid", fgColor=row_colour)

        for col, key in enumerate(RECORD_KEYS, 1):
            val = rec.get(key, "")
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Column widths
    col_widths = [22, 28, 18, 18, 14, 14, 14, 20, 14, 35, 14, 35, 20, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Status"
    ws2["B1"] = "Count"
    for i, s in enumerate(STATUSES, 2):
        ws2.cell(row=i, column=1, value=s)
        ws2.cell(row=i, column=2, value=f'=COUNTIF({SHEET_NAME}!H:H,"{s}")')
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    ws2["A12"] = "Total Applications"
    ws2["B12"] = f'=COUNTA({SHEET_NAME}!A:A)-1'

    wb.save(path)
    print(f"✅ Saved tracker → {path}  ({len(records)} records)")


# ─────────────────────────────────────────
# STEP 5: SYNC TO GOOGLE CALENDAR
# ─────────────────────────────────────────

CALENDAR_TRIGGER_STATUSES = {
    "First Round Interview", "Second Round Interview",
    "Assessment Centre", "Offer"
}

def sync_calendar_events(records: list[dict]) -> list[dict]:
    """Create Google Calendar events for interviews and offers."""
    print("📅 Syncing calendar events...")

    events_to_create = [
        (i, r) for i, r in enumerate(records)
        if r.get("status") in CALENDAR_TRIGGER_STATUSES
        and not r.get("calendar_event_id")
        and r.get("next_deadline")
    ]

    if not events_to_create:
        print("   No new calendar events needed.")
        return records

    service = get_calendar_service()
    for idx, rec in events_to_create:
        duration_hours = 4 if rec.get("status") == "Assessment Centre" else 1
        start = rec["next_deadline"]  # ISO date string e.g. "2025-03-20"
        # Build a datetime event at 9am
        start_dt = f"{start}T09:00:00"
        end_dt = f"{start}T{9 + duration_hours:02d}:00:00"

        event_body = {
            "summary": f"{rec['company']} — {rec['status']}",
            "description": (
                f"Role: {rec.get('role', '')}\n"
                f"Next step: {rec.get('next_step', '')}\n"
                f"Notes: {rec.get('notes', '')}"
            ),
            "start": {"dateTime": start_dt, "timeZone": TIMEZONE},
            "end": {"dateTime": end_dt, "timeZone": TIMEZONE},
        }

        try:
            created = service.events().insert(calendarId=CALENDAR_ID, body=event_body).execute()
            records[idx]["calendar_event_id"] = created.get("id", "")
            print(f"   📅 Created event: {event_body['summary']} on {start}")
        except HttpError as e:
            print(f"   ⚠️  Calendar event failed for {rec.get('company')}: {e}")

    return records


# ─────────────────────────────────────────
# STEP 5b: WEB SEARCH — infer rejections from offer news
# ─────────────────────────────────────────

PENDING_STATUSES = {"Applied", "Online Assessment", "Pending"}

def infer_rejections_from_web(records: list[dict]) -> list[dict]:
    """For records still in a pending status, web-search whether the company
    has already sent out offers for this cycle. If yes, mark as Rejected."""
    today = datetime.now().strftime("%Y-%m-%d")
    cycle_year = datetime.now().year

    candidates = [(i, r) for i, r in enumerate(records) if r.get("status") in PENDING_STATUSES]
    if not candidates:
        return records

    print(f"🔍 Web-searching offer status for {len(candidates)} pending application(s)...")

    for idx, rec in candidates:
        company = rec.get("company", "")
        role = rec.get("role", "")
        query = f"{company} {role} {cycle_year} offers sent spring week internship results site:reddit.com OR site:thestudentroom.co.uk OR site:efinancialcareers.com"

        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=400,
                tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 2}],
                messages=[{
                    "role": "user",
                    "content": (
                        f"Today is {today}. Search for: has {company} already sent out offers or rejections "
                        f"for their '{role}' spring week / internship programme for the {cycle_year} cycle? "
                        f"Answer YES or NO and one sentence of evidence. Be concise."
                    )
                }]
            )
            # Extract text from response
            answer = " ".join(
                block.text for block in response.content if hasattr(block, "text")
            ).strip()

            print(f"   {company}: {answer[:120]}")

            if answer.upper().startswith("YES") or "offers have been sent" in answer.lower() or "already sent" in answer.lower():
                records[idx]["status"] = "Rejected"
                records[idx]["notes"] = (
                    (records[idx].get("notes") or "") +
                    f" | Web-inferred rejection: {answer[:200]}"
                ).strip(" |")
                records[idx]["last_updated"] = today
                print(f"   → Marked {company} as Rejected (offers already out)")

        except Exception as e:
            print(f"   ⚠️  Web search failed for {company}: {e}")

    return records


# ─────────────────────────────────────────
# UPDATE MODE: load existing xlsx
# ─────────────────────────────────────────

def load_existing_records(path: str = OUTPUT_XLSX) -> tuple[list[dict], str]:
    """Load existing records from xlsx and return (records, latest_date)."""
    if not os.path.exists(path):
        return [], DATE_CUTOFF

    wb = openpyxl.load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        return [], DATE_CUTOFF

    ws = wb[SHEET_NAME]
    records = []
    latest_date = DATE_CUTOFF

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(RECORD_KEYS, row))
        records.append(rec)
        last_updated = str(rec.get("last_updated", "") or "")
        if last_updated and last_updated > latest_date:
            latest_date = last_updated[:10].replace("-", "/")

    return records, latest_date


# ─────────────────────────────────────────
# ENTRYPOINT
# ─────────────────────────────────────────

def run(update: bool = False):
    print("🚀 Starting internship tracker pipeline\n")

    existing_records = []
    date_cutoff = DATE_CUTOFF

    if update:
        existing_records, date_cutoff = load_existing_records()
        print(f"   Update mode: fetching emails after {date_cutoff}\n")

    # 1. Fetch
    threads = fetch_application_emails(date_cutoff)
    print(f"   Found {len(threads)} email threads\n")

    # 2. Parse (batched — BATCH_SIZE threads per API call)
    new_records = parse_all_threads(threads)

    # 3. Merge with existing + deduplicate
    all_records = existing_records + new_records
    records = deduplicate(all_records)
    print(f"\n   → {len(records)} unique applications after deduplication\n")

    # 4. Write xlsx
    write_xlsx(records)

    # 5. Web-search offer status for pending applications
    records = infer_rejections_from_web(records)

    # 6. Calendar sync
    records = sync_calendar_events(records)

    # 7. Re-write with calendar IDs and web inferences populated
    write_xlsx(records)

    print("\n✅ Pipeline complete.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Internship application tracker")
    parser.add_argument("--update", action="store_true",
                        help="Only fetch emails newer than the latest date in existing xlsx")
    args = parser.parse_args()
    run(update=args.update)
